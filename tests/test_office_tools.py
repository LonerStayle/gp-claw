import csv

import pytest

from gp_claw.tools.office_file import create_office_tools


@pytest.fixture
def tools(workspace):
    return create_office_tools(str(workspace))


@pytest.fixture
def excel_write(tools):
    return tools[0]


@pytest.fixture
def csv_write_tool(tools):
    return tools[1]


@pytest.fixture
def pdf_write(tools):
    return tools[2]


@pytest.fixture
def pptx_write(tools):
    return tools[3]


# --- excel_write ---


def test_excel_write_creates_file(workspace, excel_write):
    result = excel_write.invoke({
        "path": "report.xlsx",
        "sheets": [
            {
                "name": "매출",
                "headers": ["월", "금액"],
                "rows": [["1월", 1000], ["2월", 2000]],
            }
        ],
    })
    assert result["action"] == "created"
    assert result["sheets"] == 1
    assert (workspace / "report.xlsx").exists()
    assert result["size_bytes"] > 0


def test_excel_write_content_verified(workspace, excel_write):
    import openpyxl

    excel_write.invoke({
        "path": "data.xlsx",
        "sheets": [
            {
                "name": "Sheet1",
                "headers": ["이름", "나이"],
                "rows": [["홍길동", 30], ["김철수", 25]],
            }
        ],
    })
    wb = openpyxl.load_workbook(workspace / "data.xlsx")
    ws = wb.active
    assert ws.cell(1, 1).value == "이름"
    assert ws.cell(2, 1).value == "홍길동"
    assert ws.cell(3, 2).value == 25


def test_excel_write_multiple_sheets(workspace, excel_write):
    result = excel_write.invoke({
        "path": "multi.xlsx",
        "sheets": [
            {"name": "A", "headers": ["x"], "rows": [["1"]]},
            {"name": "B", "headers": ["y"], "rows": [["2"]]},
        ],
    })
    assert result["sheets"] == 2


def test_excel_write_blocks_outside_workspace(workspace, excel_write):
    with pytest.raises(Exception):
        excel_write.invoke({
            "path": "/etc/hacked.xlsx",
            "sheets": [{"name": "S", "headers": [], "rows": []}],
        })


# --- csv_write ---


def test_csv_write_creates_file(workspace, csv_write_tool):
    result = csv_write_tool.invoke({
        "path": "data.csv",
        "headers": ["이름", "점수"],
        "rows": [["Alice", 90], ["Bob", 85]],
    })
    assert result["action"] == "created"
    assert result["rows"] == 2
    assert (workspace / "data.csv").exists()


def test_csv_write_content_verified(workspace, csv_write_tool):
    csv_write_tool.invoke({
        "path": "check.csv",
        "headers": ["a", "b"],
        "rows": [["1", "2"], ["3", "4"]],
    })
    with open(workspace / "check.csv", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    assert rows[0] == ["a", "b"]
    assert rows[1] == ["1", "2"]


def test_csv_write_blocks_outside_workspace(workspace, csv_write_tool):
    with pytest.raises(Exception):
        csv_write_tool.invoke({
            "path": "/etc/hacked.csv",
            "headers": ["x"],
            "rows": [],
        })


# --- pdf_write ---


def test_pdf_write_creates_file(workspace, pdf_write):
    result = pdf_write.invoke({
        "path": "doc.pdf",
        "title": "Test Report",
        "content": "This is a test document.\nWith multiple paragraphs.",
    })
    assert result["action"] == "created"
    assert (workspace / "doc.pdf").exists()
    assert result["size_bytes"] > 0


def test_pdf_write_blocks_outside_workspace(workspace, pdf_write):
    with pytest.raises(Exception):
        pdf_write.invoke({
            "path": "/etc/hacked.pdf",
            "title": "Bad",
            "content": "bad",
        })


# --- pptx_write ---


def test_pptx_write_creates_file(workspace, pptx_write):
    result = pptx_write.invoke({
        "path": "pres.pptx",
        "title": "발표자료",
        "slides": [
            {"title": "개요", "content": "프로젝트 소개"},
            {"title": "결론", "content": "감사합니다"},
        ],
    })
    assert result["action"] == "created"
    assert result["slides"] == 3  # 타이틀 + 2 콘텐츠
    assert (workspace / "pres.pptx").exists()


def test_pptx_write_content_verified(workspace, pptx_write):
    from pptx import Presentation

    pptx_write.invoke({
        "path": "verify.pptx",
        "title": "Test",
        "slides": [{"title": "Slide1", "content": "Hello"}],
    })
    prs = Presentation(str(workspace / "verify.pptx"))
    assert len(prs.slides) == 2  # 타이틀 + 1 콘텐츠


def test_pptx_write_blocks_outside_workspace(workspace, pptx_write):
    with pytest.raises(Exception):
        pptx_write.invoke({
            "path": "/etc/hacked.pptx",
            "title": "Bad",
            "slides": [],
        })


# --- Registry integration ---


def test_office_tools_in_registry(workspace):
    from gp_claw.tools import create_tool_registry

    registry = create_tool_registry(str(workspace))
    tool_names = [t.name for t in registry.all_tools]
    assert "excel_write" in tool_names
    assert "csv_write" in tool_names
    assert "pdf_write" in tool_names
    assert "pptx_write" in tool_names


def test_office_tools_classified_as_dangerous(workspace):
    from gp_claw.tools import ToolSafety, create_tool_registry

    registry = create_tool_registry(str(workspace))
    for name in ["excel_write", "csv_write", "pdf_write", "pptx_write"]:
        assert registry.classify(name) == ToolSafety.DANGEROUS
